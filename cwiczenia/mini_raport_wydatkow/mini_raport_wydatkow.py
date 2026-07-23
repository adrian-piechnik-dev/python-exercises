import csv

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def zadanie_01_wczytaj_wydatki(sciezka: str) -> list[dict[str, str]] | None:
    """Wczytuje wydatki z pliku CSV do listy słowników.

    Args:
        sciezka: ścieżka do pliku CSV z kolumnami data, kategoria, opis, kwota.

    Returns:
        list[dict[str, str]] | None: lista wierszy jako słowniki (wszystkie
            wartości to stringi) albo None, gdy plik nie istnieje.
    """
    try:
        with open(sciezka, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return list(reader)
    except FileNotFoundError:
        return None


def zadanie_02_waliduj_wiersze(
    wiersze: list[dict[str, str]],
) -> list[dict[str, str | float]]:
    """Filtruje wiersze wydatków, zostawiając tylko te z poprawną kwotą.

    Args:
        wiersze: lista słowników z kluczem "kwota" jako string.

    Returns:
        list[dict[str, str | float]]: tylko wiersze, w których kwota dała
            się skonwertować na float i jest większa od zera; w zwróconych
            wierszach kwota jest już floatem.
    """
    poprawne = []
    for wiersz in wiersze:
        try:
            kwota = float(wiersz["kwota"])
        except ValueError:
            continue
        if kwota > 0:
            wiersz["kwota"] = kwota
            poprawne.append(wiersz)
    return poprawne


def zadanie_03_zbuduj_dataframe(
    wiersze: list[dict[str, str | float]],
) -> pd.DataFrame:
    """Buduje DataFrame z listy zwalidowanych wierszy wydatków.

    Args:
        wiersze: lista słowników po walidacji (kwota jako float).

    Returns:
        pd.DataFrame: tabela z kolumnami jak klucze słowników; kolumna
            kwota jest liczbowa, bo walidacja zrobiła konwersję wcześniej.
    """
    return pd.DataFrame(wiersze)


def zadanie_04_wydatki_powyzej(df: pd.DataFrame, prog: float) -> pd.DataFrame:
    """Zwraca wydatki o kwocie większej niż podany próg.

    Args:
        df: DataFrame z kolumną "kwota".
        prog: próg kwotowy (wydatki dokładnie równe progowi odpadają).

    Returns:
        pd.DataFrame: nowa tabela tylko z wierszami powyżej progu;
            oryginalny df pozostaje niezmieniony.
    """
    return df[df["kwota"] > prog]


def zadanie_05_suma_calkowita(df: pd.DataFrame) -> float:
    """Liczy sumę wszystkich wydatków.

    Args:
        df: DataFrame z kolumną "kwota".

    Returns:
        float: suma kolumny kwota; 0.0 dla pustej tabeli.
    """
    return float(df["kwota"].sum())


def zadanie_06_agreguj_kategorie(df: pd.DataFrame) -> pd.DataFrame:
    """Agreguje wydatki po kategorii: suma, średnia i liczba pozycji.

    Args:
        df: DataFrame z kolumnami "kategoria" i "kwota".

    Returns:
        pd.DataFrame: tabela z kolumnami kategoria, suma, srednia, liczba —
            po jednym wierszu na kategorię; kategoria jest zwykłą kolumną,
            nie indeksem.
    """
    return df.groupby("kategoria").agg(
        suma=("kwota", "sum"),
        srednia=("kwota", "mean"),
        liczba=("kwota", "count"),
    ).reset_index()


def zadanie_07_dodaj_procent(df_raport: pd.DataFrame) -> pd.DataFrame:
    """Dodaje kolumnę procentowego udziału kategorii w sumie wydatków.

    Args:
        df_raport: zagregowany raport z kolumną "suma".

    Returns:
        pd.DataFrame: nowa tabela z dodatkową kolumną "procent" —
            udział sumy kategorii w sumie całkowitej, w procentach,
            zaokrąglony do 1 miejsca; oryginał pozostaje niezmieniony.
    """
    suma_wszystkich = df_raport["suma"].sum()
    return df_raport.assign(
        procent=(df_raport["suma"] / suma_wszystkich * 100).round(1)
    )


def zadanie_08_posortuj_raport(df_raport: pd.DataFrame) -> pd.DataFrame:
    """Sortuje raport malejąco po kolumnie suma.

    Args:
        df_raport: zagregowany raport z kolumną "suma".

    Returns:
        pd.DataFrame: nowa tabela z wierszami od największej do najmniejszej
            sumy; oryginał pozostaje niezmieniony.
    """
    return df_raport.sort_values("suma", ascending=False)


def zadanie_09_eksportuj_raport(df_raport: pd.DataFrame, sciezka: str) -> bool:
    """Zapisuje raport do pliku Excel bez kolumny indeksu.

    Args:
        df_raport: gotowy raport do zapisania.
        sciezka: ścieżka do pliku wynikowego .xlsx.

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    df_raport.to_excel(sciezka, index=False)
    return True


def zadanie_10_formatuj_naglowki(sciezka: str) -> bool:
    """Formatuje wiersz nagłówków raportu: pogrubienie, szare tło, wyśrodkowanie.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx z nagłówkami w wierszu 1.

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    for komorka in ws[1]:
        komorka.font = Font(bold=True)
        komorka.fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )
        komorka.alignment = Alignment(horizontal="center")
    wb.save(sciezka)
    return True


def zadanie_11_dopasuj_uklad(sciezka: str) -> bool:
    """Ustawia szerokości kolumn raportu i zamraża wiersz nagłówków.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx z raportem.

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A2"
    wb.save(sciezka)
    return True


def zadanie_12_wyroznij_duze_wydatki(sciezka: str, prog: float) -> bool:
    """Wyróżnia czerwonym tłem komórki sum przekraczających próg.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx; sumy kategorii
            stoją w kolumnie B, dane zaczynają się od wiersza 2.
        prog: próg kwotowy (wyróżniamy tylko sumy większe od progu).

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row):
        komorka = row[0]
        if komorka.value is not None and komorka.value > prog:
            komorka.fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
    wb.save(sciezka)
    return True


def zadanie_13_generuj_raport(sciezka_csv: str, sciezka_xlsx: str) -> bool | None:
    """Buduje kompletny raport wydatków: od pliku CSV do sformatowanego Excela.

    Args:
        sciezka_csv: ścieżka do wejściowego pliku CSV z wydatkami.
        sciezka_xlsx: ścieżka do wynikowego pliku raportu .xlsx.

    Returns:
        bool | None: True po zapisaniu gotowego raportu; None, gdy plik CSV
            nie istnieje (wtedy plik raportu w ogóle nie powstaje).
    """
    wiersze = zadanie_01_wczytaj_wydatki(sciezka_csv)
    if wiersze is None:
        return None
    poprawne = zadanie_02_waliduj_wiersze(wiersze)
    df = zadanie_03_zbuduj_dataframe(poprawne)
    df_raport = zadanie_06_agreguj_kategorie(df)
    df_raport = zadanie_07_dodaj_procent(df_raport)
    df_raport = zadanie_08_posortuj_raport(df_raport)
    zadanie_09_eksportuj_raport(df_raport, sciezka_xlsx)
    zadanie_10_formatuj_naglowki(sciezka_xlsx)
    zadanie_11_dopasuj_uklad(sciezka_xlsx)
    return True
