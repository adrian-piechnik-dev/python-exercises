from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from mini_raport_wydatkow import (
    zadanie_01_wczytaj_wydatki,
    zadanie_02_waliduj_wiersze,
    zadanie_03_zbuduj_dataframe,
    zadanie_04_wydatki_powyzej,
    zadanie_05_suma_calkowita,
    zadanie_06_agreguj_kategorie,
    zadanie_07_dodaj_procent,
    zadanie_08_posortuj_raport,
    zadanie_09_eksportuj_raport,
    zadanie_10_formatuj_naglowki,
    zadanie_11_dopasuj_uklad,
    zadanie_12_wyroznij_duze_wydatki,
    zadanie_13_generuj_raport,
)


# --- zadanie_01 ---

def test_zadanie_01_wczytuje_wszystkie_wiersze(wydatki_csv: Path) -> None:
    """Co testuje: wczytanie pełnego pliku CSV do listy słowników.
    Co udaje: nic — prawdziwy plik tymczasowy z fixture wydatki_csv.
    Co sprawdzam: 7 wierszy; w pierwszym kategoria "jedzenie"
    i kwota "120.50" jako STRING (DictReader nie konwertuje).
    """
    wynik = zadanie_01_wczytaj_wydatki(str(wydatki_csv))
    assert len(wynik) == 7
    assert wynik[0]["kategoria"] == "jedzenie"
    assert wynik[0]["kwota"] == "120.50"


def test_zadanie_01_zwraca_none_gdy_brak_pliku(tmp_path: Path) -> None:
    """Co testuje: kontrakt None przy nieistniejącym pliku.
    Co udaje: nic — ścieżka w pustym katalogu tymczasowym na pewno nie istnieje.
    Co sprawdzam: wynik is None (bez wyjątku).
    """
    p = tmp_path / "nieistniejacy.csv"
    wynik = zadanie_01_wczytaj_wydatki(str(p))
    assert wynik is None


# --- zadanie_02 ---

def test_zadanie_02_odrzuca_zepsute_kwoty(
    brudne_wiersze: list[dict[str, str]],
) -> None:
    """Co testuje: bramkę jakości — zostają tylko wiersze z poprawną, dodatnią kwotą.
    Co udaje: nic — gotowa lista z fixture brudne_wiersze (2 poprawne z 5).
    Co sprawdzam: zostały dokładnie 2 wiersze, a ich kwoty to floaty 18.0 i 150.0.
    """
    wynik = zadanie_02_waliduj_wiersze(brudne_wiersze)
    assert len(wynik) == 2
    assert wynik[0]["kwota"] == 18.0
    assert wynik[1]["kwota"] == 150.0


def test_zadanie_02_pusta_lista_daje_pusta_liste() -> None:
    """Co testuje: zachowanie brzegowe dla pustego wejścia.
    Co udaje: nic — podaję pustą listę wprost.
    Co sprawdzam: wynik to pusta lista (nie None, nie wyjątek).
    """
    wynik = zadanie_02_waliduj_wiersze([])
    assert wynik == []


# --- zadanie_03 ---

def test_zadanie_03_buduje_dataframe_z_wierszy() -> None:
    """Co testuje: zamianę listy słowników na DataFrame.
    Co udaje: nic — podaję wprost małą listę 2 zwalidowanych wierszy.
    Co sprawdzam: wynik ma 2 wiersze i kolumny kategoria oraz kwota.
    """
    lista = [
        {"kategoria": "jedzenie", "kwota": 18.00},
        {"kategoria": "transport", "kwota": 150.00}
    ]
    wynik = zadanie_03_zbuduj_dataframe(lista)
    assert len(wynik) == 2
    assert "kategoria" in wynik.columns
    assert "kwota" in wynik.columns


def test_zadanie_03_kolumna_kwota_jest_liczbowa() -> None:
    """Co testuje: czy na kolumnie kwota da się od razu liczyć.
    Co udaje: nic — własna lista 2 wierszy z kwotami float.
    Co sprawdzam: suma kolumny kwota zgadza się z sumą podanych wartości.
    """
    lista = [
        {"kwota": 18.00},
        {"kwota": 150.00}
    ]
    wynik = zadanie_03_zbuduj_dataframe(lista)
    assert wynik["kwota"].sum() == 168.00


# --- zadanie_04 ---

def test_zadanie_04_filtruje_powyzej_progu(df_wydatki: pd.DataFrame) -> None:
    """Co testuje: filtr boolean z progiem.
    Co udaje: nic — gotowy DataFrame z fixture df_wydatki.
    Co sprawdzam: dla progu 100 zostają 4 wiersze (120.50, 110.00, 180.00, 150.00).
    """
    wynik = zadanie_04_wydatki_powyzej(df_wydatki, 100)
    assert len(wynik) == 4


def test_zadanie_04_nie_modyfikuje_oryginalu(df_wydatki: pd.DataFrame) -> None:
    """Co testuje: brak side effects — oryginalna tabela zostaje nietknięta.
    Co udaje: nic — fixture df_wydatki.
    Co sprawdzam: po wywołaniu funkcji oryginał nadal ma 7 wierszy.
    """
    oryginal = len(df_wydatki)
    zadanie_04_wydatki_powyzej(df_wydatki, 100)
    assert len(df_wydatki) == oryginal


# --- zadanie_05 ---

def test_zadanie_05_liczy_sume_wydatkow(df_wydatki: pd.DataFrame) -> None:
    """Co testuje: sumę całkowitą kolumny kwota.
    Co udaje: nic — fixture df_wydatki (suma kontrolna 739.0).
    Co sprawdzam: wynik == 739.0 i jest typu float.
    """
    wynik = zadanie_05_suma_calkowita(df_wydatki)
    assert isinstance(wynik, float)
    assert wynik == 739.0


def test_zadanie_05_pusta_tabela_daje_zero() -> None:
    """Co testuje: zachowanie brzegowe dla tabeli bez wierszy.
    Co udaje: nic — buduję pusty DataFrame z samą kolumną kwota.
    Co sprawdzam: wynik == 0.0.
    """
    df = pd.DataFrame({"kwota": []})
    wynik = zadanie_05_suma_calkowita(df)
    assert wynik == 0.0


# --- zadanie_06 ---

def test_zadanie_06_liczy_sumy_kategorii(df_wydatki: pd.DataFrame) -> None:
    """Co testuje: agregację sum po kategoriach.
    Co udaje: nic — fixture df_wydatki.
    Co sprawdzam: 3 wiersze wyniku; suma dla jedzenia to 261.0.
    """
    wynik = zadanie_06_agreguj_kategorie(df_wydatki)
    assert len(wynik) == 3
    suma_jedzenie = wynik[wynik["kategoria"] == "jedzenie"]["suma"].tolist()[0]
    assert suma_jedzenie == 261.0


def test_zadanie_06_ma_wszystkie_kolumny_raportu(
    df_wydatki: pd.DataFrame,
) -> None:
    """Co testuje: nazwy kolumn z nazwanej agregacji + reset_index.
    Co udaje: nic — fixture df_wydatki.
    Co sprawdzam: wynik ma kolumny kategoria, suma, srednia, liczba;
    liczba dla transportu to 2.
    """
    wynik = zadanie_06_agreguj_kategorie(df_wydatki)
    assert wynik.columns.tolist() == ["kategoria", "suma", "srednia", "liczba"]
    liczba_transport = (
        wynik[wynik["kategoria"] == "transport"]["liczba"].tolist()[0]
    )
    assert liczba_transport == 2


# --- zadanie_07 ---

def test_zadanie_07_liczy_udzial_procentowy(df_raport: pd.DataFrame) -> None:
    """Co testuje: kolumnę procent (udział sumy kategorii w całości).
    Co udaje: nic — fixture df_raport (całość 739.0).
    Co sprawdzam: procent jedzenia == 35.3 (261/739*100 po zaokrągleniu).
    """
    wynik = zadanie_07_dodaj_procent(df_raport)
    assert "procent" in wynik.columns
    assert wynik[wynik["kategoria"] == "jedzenie"]["procent"].tolist()[0] == 35.3


def test_zadanie_07_nie_modyfikuje_oryginalu(df_raport: pd.DataFrame) -> None:
    """Co testuje: brak side effects — fixture nie dostaje nowej kolumny.
    Co udaje: nic — fixture df_raport.
    Co sprawdzam: po wywołaniu w oryginale NADAL nie ma kolumny procent.
    """
    zadanie_07_dodaj_procent(df_raport)
    assert "procent" not in df_raport


# --- zadanie_08 ---

def test_zadanie_08_najwieksza_suma_na_gorze(df_raport: pd.DataFrame) -> None:
    """Co testuje: sortowanie malejąco po sumie.
    Co udaje: nic — fixture df_raport (największa suma: transport 290.0).
    Co sprawdzam: kolejność kategorii to transport, jedzenie, rozrywka.
    """
    wynik = zadanie_08_posortuj_raport(df_raport)
    assert wynik["kategoria"].tolist() == ["transport", "jedzenie", "rozrywka"]


def test_zadanie_08_nie_modyfikuje_oryginalu(df_raport: pd.DataFrame) -> None:
    """Co testuje: brak side effects przy sortowaniu.
    Co udaje: nic — fixture df_raport (pierwszy wiersz: jedzenie).
    Co sprawdzam: po wywołaniu pierwszy wiersz oryginału to nadal jedzenie.
    """
    zadanie_08_posortuj_raport(df_raport)
    assert df_raport["kategoria"].tolist()[0] == "jedzenie"


# --- zadanie_09 ---

def test_zadanie_09_tworzy_plik_i_zwraca_true(
    df_raport: pd.DataFrame, tmp_path: Path,
) -> None:
    """Co testuje: eksport raportu do pliku .xlsx.
    Co udaje: nic — zapis do prawdziwego pliku w tmp_path.
    Co sprawdzam: wynik is True i plik istnieje na dysku.
    """
    sciezka = tmp_path / "plik.xlsx"
    wynik = zadanie_09_eksportuj_raport(df_raport, str(sciezka))
    assert wynik is True
    assert Path(sciezka).exists()


def test_zadanie_09_bez_kolumny_indeksu(
    df_raport: pd.DataFrame, tmp_path: Path,
) -> None:
    """Co testuje: czy eksport pominął indeks (index=False).
    Co udaje: nic — zapis i odczyt prawdziwego pliku.
    Co sprawdzam: komórka A1 zawiera "kategoria" (a nie pusty nagłówek indeksu).
    """
    sciezka = tmp_path / "plik.xlsx"
    zadanie_09_eksportuj_raport(df_raport, str(sciezka))
    wb = load_workbook(str(sciezka))
    ws = wb.active
    assert ws["A1"].value == "kategoria"


# --- zadanie_10 ---

def test_zadanie_10_pogrubia_wszystkie_naglowki(raport_xlsx: Path) -> None:
    """Co testuje: pogrubienie całego wiersza nagłówków.
    Co udaje: nic — gotowy plik z fixture raport_xlsx (nagłówki A1-D1).
    Co sprawdzam: czcionka A1 i D1 ma bold ustawione na True.
    """
    wynik = zadanie_10_formatuj_naglowki(str(raport_xlsx))
    wb = load_workbook(raport_xlsx)
    ws = wb.active
    assert ws["A1"].font.bold is True
    assert ws["D1"].font.bold is True
    assert wynik is True


def test_zadanie_10_tlo_i_wysrodkowanie(raport_xlsx: Path) -> None:
    """Co testuje: wypełnienie tła i wyśrodkowanie nagłówków.
    Co udaje: nic — fixture raport_xlsx.
    Co sprawdzam: A1 ma fill typu solid i wyrównanie poziome center.
    """
    zadanie_10_formatuj_naglowki(str(raport_xlsx))
    wb = load_workbook(raport_xlsx)
    ws = wb.active
    assert ws["A1"].fill.patternType == "solid"
    assert ws["A1"].alignment.horizontal == "center"


# --- zadanie_11 ---

def test_zadanie_11_ustawia_szerokosci_kolumn(raport_xlsx: Path) -> None:
    """Co testuje: szerokości kolumn raportu.
    Co udaje: nic — fixture raport_xlsx.
    Co sprawdzam: kolumna A ma szerokość 18, kolumna B ma 12.
    """
    zadanie_11_dopasuj_uklad(str(raport_xlsx))
    wb = load_workbook(str(raport_xlsx))
    ws = wb.active
    assert ws.column_dimensions["A"].width == 18
    assert ws.column_dimensions["B"].width == 12


def test_zadanie_11_zamraza_wiersz_naglowkow(raport_xlsx: Path) -> None:
    """Co testuje: zamrożenie pierwszego wiersza.
    Co udaje: nic — fixture raport_xlsx.
    Co sprawdzam: freeze_panes arkusza to "A2".
    """
    zadanie_11_dopasuj_uklad(str(raport_xlsx))
    wb = load_workbook(str(raport_xlsx))
    ws = wb.active
    assert ws.freeze_panes == "A2"


# --- zadanie_12 ---

def test_zadanie_12_wyroznia_sumy_powyzej_progu(raport_xlsx: Path) -> None:
    """Co testuje: czerwone tło tylko dla sum powyżej progu.
    Co udaje: nic — fixture raport_xlsx (sumy: B2=261, B3=290, B4=188).
    Co sprawdzam: dla progu 250 komórki B2 i B3 mają fill typu solid,
    a B4 nie ma (patternType is None).
    """
    wynik = zadanie_12_wyroznij_duze_wydatki(str(raport_xlsx), 250)
    assert wynik is True
    wb = load_workbook(raport_xlsx)
    ws = wb.active
    assert ws["B2"].fill.patternType == "solid"
    assert ws["B3"].fill.patternType == "solid"
    assert ws["B4"].fill.patternType is None


def test_zadanie_12_prog_wyzszy_niz_wszystko_nic_nie_barwi(
    raport_xlsx: Path,
) -> None:
    """Co testuje: zachowanie brzegowe — próg nieosiągalny.
    Co udaje: nic — fixture raport_xlsx.
    Co sprawdzam: dla progu 1000 komórka B3 (największa suma) NIE ma
    wypełnienia solid.
    """
    zadanie_12_wyroznij_duze_wydatki(str(raport_xlsx), 1000)
    wb = load_workbook(raport_xlsx)
    ws = wb.active
    assert ws["B3"].fill.patternType is None


# --- zadanie_13 ---

def test_zadanie_13_buduje_kompletny_raport(
    wydatki_csv: Path, tmp_path: Path,
) -> None:
    """Co testuje: cały pipeline od CSV do sformatowanego Excela.
    Co udaje: nic — prawdziwy CSV z fixture, zapis do tmp_path.
    Co sprawdzam: wynik is True; w pliku A1 == "kategoria",
    a B2 == 290.0 (transport na górze po sortowaniu malejąco).
    """
    sciezka = tmp_path / "raport.xlsx"
    wynik = zadanie_13_generuj_raport(str(wydatki_csv), str(sciezka))
    assert wynik is True
    wb = load_workbook(str(sciezka))
    ws = wb.active
    assert ws["A1"].value == "kategoria"
    assert ws["B2"].value == 290.0


def test_zadanie_13_none_gdy_brak_csv(tmp_path: Path) -> None:
    """Co testuje: propagację kontraktu None przez cały pipeline.
    Co udaje: nic — ścieżka CSV, która nie istnieje.
    Co sprawdzam: wynik is None i plik raportu NIE powstał.
    """
    sciezka_csv = tmp_path / "brak.csv"
    sciezka_xlsx = tmp_path / "raport.xlsx"
    wynik = zadanie_13_generuj_raport(str(sciezka_csv), str(sciezka_xlsx))
    assert wynik is None
    assert not sciezka_xlsx.exists()
