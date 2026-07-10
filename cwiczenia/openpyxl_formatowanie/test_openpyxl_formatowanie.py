from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from openpyxl_formatowanie import (
    zadanie_01_utworz_i_zapisz,
    zadanie_02_wczytaj_komorke,
    zadanie_03_pogrub_naglowki,
    zadanie_04_koloruj_czcionke,
    zadanie_05_wypelnij_tlo,
    zadanie_06_wysrodkuj_naglowki,
    zadanie_07_dodaj_obramowanie,
    zadanie_08_zamroz_naglowek,
    zadanie_09_ustaw_szerokosc,
    zadanie_10_zsumuj_kolumne_b,
    zadanie_11_zapisz_dataframe,
    zadanie_12_raport_sprzedazy,
)


# --- zadanie_01 ---

def test_zadanie_01_tworzy_plik_na_dysku(tmp_path: Path) -> None:
    """Co testuje: czy po wywołaniu funkcji plik .xlsx istnieje na dysku.
    Co udaje: nic — używam tmp_path jako cel zapisu.
    Co sprawdzam: ścieżka wskazuje istniejący plik (p.exists() is True).
    """
    p = tmp_path / "nowy.xlsx"
    zadanie_01_utworz_i_zapisz(str(p), "Raport")
    assert p.exists() is True


def test_zadanie_01_zapisuje_wartosc_i_zwraca_true(tmp_path: Path) -> None:
    """Co testuje: czy wartość trafia do A1 i funkcja zwraca True.
    Co udaje: nic — używam tmp_path jako cel zapisu.
    Co sprawdzam: wynik is True; po otwarciu pliku ws["A1"].value == "Raport".
    """
    p = tmp_path / "nowy.xlsx"
    wynik = zadanie_01_utworz_i_zapisz(str(p), "Raport")
    assert wynik is True
    wb = load_workbook(str(p))
    ws = wb.active
    assert ws["A1"].value == "Raport"


# --- zadanie_02 ---

def test_zadanie_02_odczytuje_naglowek(plik_xlsx: Path) -> None:
    """Co testuje: czy funkcja zwraca wartość tekstową komórki A1.
    Co udaje: nic — używam fixture plik_xlsx (A1 zawiera "miasto").
    Co sprawdzam: wynik == "miasto".
    """
    wynik = zadanie_02_wczytaj_komorke(str(plik_xlsx), "A1")
    assert wynik == "miasto"


def test_zadanie_02_odczytuje_liczbe(plik_xlsx: Path) -> None:
    """Co testuje: czy funkcja zwraca wartość liczbową (int, nie obiekt Cell).
    Co udaje: nic — używam fixture plik_xlsx (B2 zawiera 100).
    Co sprawdzam: wynik == 100.
    """
    wynik = zadanie_02_wczytaj_komorke(str(plik_xlsx),"B2")
    assert wynik == 100


# --- zadanie_03 ---

def test_zadanie_03_naglowki_pogrubione(plik_xlsx: Path) -> None:
    """Co testuje: czy po wywołaniu funkcji A1 i B1 mają pogrubioną czcionkę.
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: po ponownym otwarciu pliku font.bold is True dla A1 i B1.
    """
    zadanie_03_pogrub_naglowki(str(plik_xlsx))
    wb = load_workbook(str(plik_xlsx))
    ws = wb.active
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True


def test_zadanie_03_zwraca_true(plik_xlsx: Path) -> None:
    """Co testuje: czy funkcja zwraca True po zapisie (kontrakt bool).
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: wynik is True.
    """
    wynik = zadanie_03_pogrub_naglowki(str(plik_xlsx))
    assert wynik is True


# --- zadanie_04 ---

def test_zadanie_04_kolor_czcionki_ustawiony(plik_xlsx: Path) -> None:
    """Co testuje: czy komórka A1 dostała czerwoną czcionkę.
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: font.color.rgb kończy się na "FF0000" (przez doklejaną alfę).
    """
    zadanie_04_koloruj_czcionke(str(plik_xlsx), "A1", "FF0000")
    wb = load_workbook(str(plik_xlsx))
    ws = wb.active
    assert ws["A1"].font.color.rgb.endswith("FF0000") is True


def test_zadanie_04_zwraca_true(plik_xlsx: Path) -> None:
    """Co testuje: czy funkcja zwraca True po zapisie (kontrakt bool).
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: wynik is True.
    """
    wynik = zadanie_04_koloruj_czcionke(str(plik_xlsx), "A1", "FF0000")
    assert wynik is True


# --- zadanie_05 ---

def test_zadanie_05_tlo_wypelnione_kolorem(plik_xlsx: Path) -> None:
    """Co testuje: czy komórka A1 dostała żółte tło z pełnym wypełnieniem.
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: fill.start_color.rgb kończy się na "FFFF00"
    i fill.fill_type == "solid".
    """
    zadanie_05_wypelnij_tlo(str(plik_xlsx), "A1", "FFFF00")
    wb = load_workbook(str(plik_xlsx))
    ws = wb.active
    assert ws["A1"].fill.start_color.rgb.endswith("FFFF00") is True
    assert ws["A1"].fill.fill_type == "solid"


def test_zadanie_05_zwraca_true(plik_xlsx: Path) -> None:
    """Co testuje: czy funkcja zwraca True po zapisie (kontrakt bool).
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: wynik is True.
    """
    wynik = zadanie_05_wypelnij_tlo(str(plik_xlsx), "B1", "FFFF00")
    assert wynik is True


# --- zadanie_06 ---

def test_zadanie_06_naglowki_wysrodkowane(plik_xlsx: Path) -> None:
    """Co testuje: czy A1 i B1 mają wyśrodkowanie poziome i pionowe.
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: alignment.horizontal == "center"
    i alignment.vertical == "center" dla obu komórek.
    """
    zadanie_06_wysrodkuj_naglowki(str(plik_xlsx))
    wb = load_workbook(str(plik_xlsx))
    ws = wb.active
    assert ws["A1"].alignment.horizontal == "center"
    assert ws["A1"].alignment.vertical == "center"
    assert ws["B1"].alignment.horizontal == "center"
    assert ws["B1"].alignment.vertical == "center"


def test_zadanie_06_zwraca_true(plik_xlsx: Path) -> None:
    """Co testuje: czy funkcja zwraca True po zapisie (kontrakt bool).
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: wynik is True.
    """
    wynik = zadanie_06_wysrodkuj_naglowki(str(plik_xlsx))
    assert wynik is True


# --- zadanie_07 ---

def test_zadanie_07_obramowanie_ze_wszystkich_stron(plik_xlsx: Path) -> None:
    """Co testuje: czy komórka A1 ma cienką ramkę z każdej z czterech stron.
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: border.left/right/top/bottom.style == "thin".
    """
    zadanie_07_dodaj_obramowanie(str(plik_xlsx), "A1")
    wb = load_workbook(str(plik_xlsx))
    ws = wb.active
    assert ws["A1"].border.left.style == "thin"
    assert ws["A1"].border.right.style == "thin"
    assert ws["A1"].border.top.style == "thin"
    assert ws["A1"].border.bottom.style == "thin"


def test_zadanie_07_zwraca_true(plik_xlsx: Path) -> None:
    """Co testuje: czy funkcja zwraca True po zapisie (kontrakt bool).
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: wynik is True.
    """
    wynik = zadanie_07_dodaj_obramowanie(str(plik_xlsx), "B2")
    assert wynik is True


# --- zadanie_08 ---

def test_zadanie_08_freeze_ustawiony_na_a2(plik_xlsx: Path) -> None:
    """Co testuje: czy arkusz ma zamrożony wiersz nagłówków.
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: ws.freeze_panes == "A2" po ponownym otwarciu pliku.
    """
    zadanie_08_zamroz_naglowek(str(plik_xlsx))
    wb = load_workbook(str(plik_xlsx))
    ws = wb.active
    assert ws.freeze_panes == "A2"


def test_zadanie_08_zwraca_true(plik_xlsx: Path) -> None:
    """Co testuje: czy funkcja zwraca True po zapisie (kontrakt bool).
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: wynik is True.
    """
    wynik = zadanie_08_zamroz_naglowek(str(plik_xlsx))
    assert wynik is True


# --- zadanie_09 ---

def test_zadanie_09_szerokosc_kolumny_ustawiona(plik_xlsx: Path) -> None:
    """Co testuje: czy kolumna A dostała szerokość 25.
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: ws.column_dimensions["A"].width == 25 po otwarciu pliku.
    """
    zadanie_09_ustaw_szerokosc(str(plik_xlsx), "A", 25)
    wb = load_workbook(str(plik_xlsx))
    ws = wb.active
    assert ws.column_dimensions["A"].width == 25


def test_zadanie_09_zwraca_true(plik_xlsx: Path) -> None:
    """Co testuje: czy funkcja zwraca True po zapisie (kontrakt bool).
    Co udaje: nic — używam fixture plik_xlsx.
    Co sprawdzam: wynik is True.
    """
    wynik = zadanie_09_ustaw_szerokosc(str(plik_xlsx), "B", 30)
    assert wynik is True


# --- zadanie_10 ---

def test_zadanie_10_sumuje_kolumne_b(plik_xlsx: Path) -> None:
    """Co testuje: czy suma kolumny B pomija nagłówek i liczy tylko dane.
    Co udaje: nic — używam fixture plik_xlsx (100 + 200 + 300 = 600).
    Co sprawdzam: wynik == 600.
    """
    wynik = zadanie_10_zsumuj_kolumne_b(str(plik_xlsx))
    assert wynik == 600


def test_zadanie_10_same_naglowki_daja_zero(tmp_path: Path) -> None:
    """Co testuje: czy plik z samymi nagłówkami (bez wierszy danych) daje 0.
    Co udaje: nic — buduję własny plik przez Workbook i tmp_path.
    Co sprawdzam: wynik == 0 (a nie błąd ani None).
    """
    p = tmp_path / "puste.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["miasto", "sprzedaz"])
    wb.save(str(p))
    wynik = zadanie_10_zsumuj_kolumne_b(str(p))
    assert wynik == 0


# --- zadanie_11 ---

def test_zadanie_11_plik_ma_naglowki_bez_indeksu(
    tmp_path: Path, df_sprzedaz: pd.DataFrame
) -> None:
    """Co testuje: czy plik zaczyna się od nagłówka "miasto" w A1 (bez kolumny indeksu).
    Co udaje: nic — używam fixture df_sprzedaz i tmp_path.
    Co sprawdzam: ws["A1"].value == "miasto" (indeks nie przesunął kolumn).
    """
    p = tmp_path / "dane.xlsx"
    zadanie_11_zapisz_dataframe(str(p), df_sprzedaz)
    wb = load_workbook(str(p))
    ws = wb.active
    assert ws["A1"].value == "miasto"


def test_zadanie_11_dwie_kolumny_i_true(
    tmp_path: Path, df_sprzedaz: pd.DataFrame
) -> None:
    """Co testuje: czy plik ma dokładnie 2 kolumny (index=False) i funkcja zwraca True.
    Co udaje: nic — używam fixture df_sprzedaz i tmp_path.
    Co sprawdzam: wynik is True i ws.max_column == 2.
    """
    p = tmp_path / "dane.xlsx"
    wynik = zadanie_11_zapisz_dataframe(str(p), df_sprzedaz)
    assert wynik is True
    wb = load_workbook(str(p))
    ws = wb.active
    assert ws.max_column == 2


# --- zadanie_12 ---

def test_zadanie_12_raport_ma_zsumowane_wartosci(
    tmp_path: Path, df_sprzedaz: pd.DataFrame
) -> None:
    """Co testuje: czy raport zawiera sprzedaż zsumowaną per miasto.
    Co udaje: nic — używam fixture df_sprzedaz (Krakow=200, Warszawa=100+300=400;
    groupby sortuje grupy alfabetycznie, więc Krakow w wierszu 2, Warszawa w 3).
    Co sprawdzam: A2=="Krakow", B2==200, A3=="Warszawa", B3==400.
    """
    p = tmp_path / "raport.xlsx"
    zadanie_12_raport_sprzedazy(str(p), df_sprzedaz)
    wb = load_workbook(str(p))
    ws = wb.active
    assert ws["A2"].value == "Krakow"
    assert ws["B2"].value == 200
    assert ws["A3"].value == "Warszawa"
    assert ws["B3"].value == 400


def test_zadanie_12_naglowek_pogrubiony_i_zamrozony(
    tmp_path: Path, df_sprzedaz: pd.DataFrame
) -> None:
    """Co testuje: czy raport ma pogrubione nagłówki, zamrożony wiersz 1
    i funkcja zwraca True.
    Co udaje: nic — używam fixture df_sprzedaz i tmp_path.
    Co sprawdzam: wynik is True, font.bold is True dla A1 i B1,
    ws.freeze_panes == "A2".
    """
    p = tmp_path / "raport.xlsx"
    wynik = zadanie_12_raport_sprzedazy(str(p), df_sprzedaz)
    assert wynik is True
    wb = load_workbook(str(p))
    ws = wb.active
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True
    assert ws.freeze_panes == "A2"