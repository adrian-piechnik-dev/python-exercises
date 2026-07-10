from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def zadanie_01_utworz_i_zapisz(sciezka: str, wartosc: str) -> bool:
    """Tworzy nowy skoroszyt, wpisuje wartość do komórki A1 i zapisuje plik.

    Args:
        sciezka: ścieżka do pliku wynikowego .xlsx.
        wartosc: tekst do wpisania w komórce A1.

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"] = wartosc
    wb.save(sciezka)
    return True


def zadanie_02_wczytaj_komorke(sciezka: str, adres: str) -> Any:
    """Wczytuje istniejący plik Excela i zwraca wartość komórki o podanym adresie.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx.
        adres: adres komórki, np. "A1" albo "B2".

    Returns:
        Any: wartość komórki (str, int lub None gdy komórka pusta).
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    return ws[adres].value


def zadanie_03_pogrub_naglowki(sciezka: str) -> bool:
    """Pogrubia czcionkę komórek nagłówka A1 i B1 w istniejącym pliku.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx z nagłówkami w wierszu 1.

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    wb.save(sciezka)
    return True


def zadanie_04_koloruj_czcionke(sciezka: str, adres: str, kolor: str) -> bool:
    """Ustawia kolor czcionki wskazanej komórki w istniejącym pliku.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx.
        adres: adres komórki, np. "A1".
        kolor: kolor w zapisie hex RRGGBB, np. "FF0000" (bez znaku #).

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    ws[adres].font = Font(color=kolor)
    wb.save(sciezka)
    return True


def zadanie_05_wypelnij_tlo(sciezka: str, adres: str, kolor: str) -> bool:
    """Wypełnia tło wskazanej komórki jednolitym kolorem.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx.
        adres: adres komórki, np. "A1".
        kolor: kolor w zapisie hex RRGGBB, np. "FFFF00".

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    ws[adres].fill = PatternFill(
        start_color=kolor,
        end_color=kolor,
        fill_type="solid"
    )
    wb.save(sciezka)
    return True


def zadanie_06_wysrodkuj_naglowki(sciezka: str) -> bool:
    """Wyśrodkowuje (w poziomie i pionie) zawartość komórek nagłówka A1 i B1.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx z nagłówkami w wierszu 1.

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    ws["B1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    wb.save(sciezka)
    return True


def zadanie_07_dodaj_obramowanie(sciezka: str, adres: str) -> bool:
    """Dodaje cienkie obramowanie wskazanej komórce ze wszystkich czterech stron.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx.
        adres: adres komórki, np. "A1".

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    cienka = Side(style="thin")
    ws[adres].border = Border(
        left=cienka,
        right=cienka,
        top=cienka,
        bottom=cienka,
    )
    wb.save(sciezka)
    return True


def zadanie_08_zamroz_naglowek(sciezka: str) -> bool:
    """Zamraża wiersz nagłówków, żeby nie uciekał przy przewijaniu.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx z nagłówkami w wierszu 1.

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    ws.freeze_panes = "A2"
    wb.save(sciezka)
    return True


def zadanie_09_ustaw_szerokosc(sciezka: str, kolumna: str, szerokosc: float) -> bool:
    """Ustawia szerokość wskazanej kolumny.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx.
        kolumna: litera kolumny, np. "A" (nie adres komórki!).
        szerokosc: szerokość kolumny w znakach, np. 25.

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    ws.column_dimensions[kolumna].width = szerokosc
    wb.save(sciezka)
    return True


def zadanie_10_zsumuj_kolumne_b(sciezka: str) -> int:
    """Sumuje wartości liczbowe z kolumny B, pomijając nagłówek z wiersza 1.

    Args:
        sciezka: ścieżka do istniejącego pliku .xlsx z nagłówkami w wierszu 1
            i liczbami w kolumnie B.

    Returns:
        int: suma wartości z kolumny B; 0 gdy plik ma tylko nagłówki.
    """
    wb = load_workbook(sciezka)
    ws = wb.active
    suma = 0
    for wiersz in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(wiersz[1], (int, float)):
            suma += wiersz[1]
    return suma


def zadanie_11_zapisz_dataframe(sciezka: str, df: pd.DataFrame) -> bool:
    """Zapisuje DataFrame do pliku Excela bez dodatkowej kolumny indeksu.

    Args:
        sciezka: ścieżka do pliku wynikowego .xlsx.
        df: tabela danych do zapisania.

    Returns:
        bool: True po pomyślnym zapisie pliku.
    """
    df.to_excel(sciezka, index=False)
    return True


def zadanie_12_raport_sprzedazy(sciezka: str, df: pd.DataFrame) -> bool:
    """Buduje raport: sumuje sprzedaż per miasto i zapisuje do Excela
    z pogrubionym, zamrożonym nagłówkiem.

    Args:
        sciezka: ścieżka do pliku wynikowego .xlsx.
        df: tabela danych z kolumnami "miasto" i "sprzedaz".

    Returns:
        bool: True po pomyślnym zapisie sformatowanego pliku.
    """
    wynik = (
        df
        .groupby("miasto")
        .agg({"sprzedaz": "sum"})
        .reset_index()
    )
    wynik.to_excel(sciezka, index=False)
    wb = load_workbook(sciezka)
    ws = wb.active
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws.freeze_panes = "A2"
    wb.save(sciezka)
    return True
